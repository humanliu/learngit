#import math
import xlrd
import xlwt

import openpyxl
from openpyxl import Workbook
from openpyxl  import load_workbook

#a = [1,2,3,4,5,6,7,8,9,10]
# print (a)

# sum=0
# for i in a:
#     sum=sum+i
# print (sum)

# Fibonacci series: 斐波纳契数列
# 两个元素的总和确定了下一个数
# a,b=0,1
# while b<1000:
#     print (b,end=',')
#     a,b=b,a+b

# xpath = r'E:\workplace\data-python' # 相对路径目录
# xlspath = r'202107考勤.xls' # 相对路径
# xxlspath = xpath + "\\" + xlspath # 绝对路径
#print (xxlspath)

# data = xlrd.open_workbook(xxlspath)
# table = data.sheets()[0]#打开第一张表
# nrows = table.nrows # 获取表的行数
# ncols = table.ncols # 获取表的列数
# worksheets = data.sheet_names() # 抓取所有sheet页的名称

# print (nrows) # 打印行数
# print (ncols) # 打印列数
# for i in range(nrows):
#    print (table.row_values(i)[:4])#取前4列数据
# for i in range(ncols):
#     print (table.col_values(i)[:4])#取前4行

# print('工作表是%s'%worksheets) # 打印工作簿名称
# worksheet1 = data.sheet_by_name(u'考勤表')

# num_rows = worksheet1.nrows # 遍历 sheet1中所有行row
# num_cols = worksheet1.ncols # 遍历sheet1中所有列col
# 遍历 sheet1中所有单元格cell
# for rown in range(num_rows):
#     for coln in range(num_cols):
#         cell = worksheet1.cell_value(rown,coln)
#         print(cell)

#xlwt模块-创建workbook和sheet对象
# workbook = xlwt.Workbook()
# sheet1 = workbook.add_sheet('sheet1',cell_overwrite_ok=True)

# sheet1.write(0,0,'name1')
# sheet1.write(0,1,'name2')
# sheet1.write(1,0,'name3')
# sheet1.write(1,1,'name4')
# sheet1.write(1,2,'name5')

# workbook.save('test-xlwt.xlsx')

# print('创建excel完成')

#openpyxl库-写入数据
xpath = r'D:\data\python' # 相对路径目录
xlsxpath = r'test.xlsx' # 相对路径
xxlsxpath = xpath + "\\" + xlsxpath # 绝对路径
#print (xxlsxpath)


wb = openpyxl.load_workbook(xxlsxpath) #获取表格
#wb = Workbook()
sheet = wb.active #选择活动的工作簿
# sheet1 = wb['sheet1']
# sheet['D10'] = 'hello world' #对某单元格赋值
sheet['D11'] = '梁一'
wb.save('test.xlsx') #保存
data = sheet['D11'].value
print(data)

# data = xlrd.open_workbook(xxlsxpath)
# table = data.sheets()[0]#打开第一张表
# worksheets = data.sheet_nam
# es() # 抓取所有sheet页的名称
# print('工作表是%s'%worksheets) # 打印工作表




